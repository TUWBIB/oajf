import io
import traceback
from functools import wraps
from typing import List, Tuple

import requests
import openpyxl
import csv

from flask import g, current_app as app, request
from flask_babel import lazy_gettext as _

from oajf.db import readPublishers as db_readPublishers
from oajf.db import readSettings as db_readSettings
from oajf.models import Journal

def logfunc(f):
    from oajf.db import getPoolStats
    @wraps(f)
    def decorated_function(*args, **kwargs):
        session_id = None
        try:
            session_id = request.cookies.get(app.config["SESSION_COOKIE_NAME"])            
        except Exception as e:
            pass
        free, used, has_pool = getPoolStats()
        if has_pool:
            app.logger.debug(f'*** {f.__name__}, free: {free}, used: {used}, session {session_id}')
        else:
            app.logger.debug(f'*** {f.__name__}, session {session_id}')
        retval = None
        try:
            retval = f(*args, **kwargs)
        except Exception as e:
            app.logger.error(e)
            app.logger.error(traceback.format_exc())
            raise e
        free, used, has_pool = getPoolStats()
        if has_pool:
            app.logger.debug(f'--- {f.__name__}, free: {free}, used: {used}, session {session_id}')
        else:
            app.logger.debug(f'--- {f.__name__}, session {session_id}')
        return retval
    return decorated_function


@logfunc
def get_publishers(force_reload=False):
    publishers = getattr(g, 'publishers', None)
    if publishers is None or force_reload:
        publishers, m_publishers = db_readPublishers()

        g.publishers = publishers
        g.m_publishers = m_publishers
        
    return publishers

@logfunc
def get_settings(force_reload=False):
    l_setting = getattr(g, 'l_setting', None)
    if l_setting is None or force_reload:
        l_setting = db_readSettings()
        g.l_setting = l_setting
        m_id_setting = {}
        m_name_setting = {}
        for o in l_setting:
            m_id_setting[o.id] = o
            m_name_setting[o.name] = o

        g.l_setting = l_setting
        g.m_id_setting = m_id_setting
        g.m_name_setting = m_name_setting
        
    return l_setting

def getSettingValue(name):
    get_settings()
    x = g.m_name_setting.get(name,None)
    return x.value if x else None

def getSettingValueLang(name,lang):
    val = None
    get_settings()
    x = g.m_name_setting.get(name,None)
    if x:
        val = getattr(x,'value_'+lang,None)
    
    return val


def getDOAJChangesFileAsExcelWorkbook(url=None) -> Tuple[openpyxl.workbook.Workbook,io.BytesIO,List[str]]:
    """
    fetches the DOAJ-changes file from Google-Docs and returns it as parsed openpyxl Workbook together with the raw bytes
    also checks the file for conformance to the expected format
    """
    errs = []
    wb = None
    data = None

    if not url:
        url = getSettingValue('doaj_changes_link')
    if not url:
        errs.append(_("URL für DOAJ-Änderungen nicht gesetzt."))
        return [],errs

    try:
        r = requests.get(url, allow_redirects=True)
    except Exception as e:
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
        errs.append(_(f"Fehler beim Holen der DOAJ-Änderungen."))
        errs.append(e)
        return wb,data,errs

    try:
        data = io.BytesIO(r.content)
        data.seek(0)
        wb = openpyxl.load_workbook(filename=data)
    except Exception as e:
        errs.append(_(f"Fehler beim Parsen der DOAJ-Änderungen."))
        errs.append(e)
        return wb,data,errs

    try:
        sheet = wb['Withdrawn']
        vals = []
        for i in range(1,5):
            val = sheet.cell(7,i).value
            if val: val = str(val).strip()
            vals.append(val)
        if (
            vals[0] != 'Journal Title' or
            vals[1] != 'ISSN' or
            vals[2] != 'Date Removed (dd/mm/yyyy)' or
            vals[3] != 'Reason'
        ):
            errs.append("sheet 'Withdrawn' does not conform to the expected format")
            return wb,data,errs
        
        sheet = wb['Added']
        vals = []
        for i in range(1,4):
            val = sheet.cell(6,i).value
            if val: val = str(val).strip()
            vals.append(val)
        if (
            vals[0] != 'Journal Title' or
            vals[1] != 'ISSN' or
            vals[2] != 'Date Added'
        ):
            errs.append(_(f"Das sheet 'Added' entspricht nicht dem erwarteten Format."))
            return wb,data,errs
    except Exception as e:
        errs.append(_(f"Fehler beimn Prüfen des Formats der DOAJ-Änderungen."))
        errs.append(e)
    
        return wb,data,errs

    return wb,data,errs


def getDOAJDump(url=None) -> Tuple[List[Journal],List[str]]:
    """
    fetches the full DOAJ dump (csv file) and returns it as a list of journals
    also checks the file for conformance to the expected format
    """
    errs = []
    l_journal: List[Journal] = []

    if not url:
        url = getSettingValue('doaj_dump_link')
    if not url:
        errs.append(_("URL für DOAJ-Dump nicht gesetzt."))
        return [],errs

    try:
        r = requests.get(url, allow_redirects=True)
    except Exception as e:
        errs.append(_(f"Fehler beim Holden des DOAJ-Dumps {e}"))
        return [],errs
    
    try:
        data = io.StringIO(r.content.decode('utf-8'),newline='')
        data.seek(0)

        reader = csv.DictReader(data, delimiter=',', quotechar='"')
        for row in reader:
            j = Journal()
            j.title = row.get('Journal title',None)
            j.url = row.get('URL in DOAJ',None)
            j.print_issn = row.get('Journal ISSN (print version)',None)
            j.e_issn = row.get('Journal EISSN (online version)',None)
            j.added_on_date = row.get('Added on Date',None)
            j.last_updated_date= row.get('Last updated Date',None)
            l_journal.append(j)
    except Exception as e:
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
        errs.append(_(f"Fehler beim Parsen des DOAJ-Dumps."))
        errs.append(e)
        return [],errs
    
    return l_journal,[]

