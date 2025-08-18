from __future__ import annotations

import os
import sys
import csv
import logging
import logging.config
import hashlib
import json
import re
import io
import ssl
import mariadb
import functools
import urllib.parse
import click
import traceback
import datetime
import openpyxl.workbook
import xlsxwriter
import atexit
import signal
import requests
import openpyxl
from json import JSONEncoder
from typing import List, Dict, BinaryIO,Tuple

from ldap3 import Server, Connection, Tls, ALL
from base64 import b64decode as decode

from flask import Flask, render_template, g, request, session, redirect, url_for, flash, send_file, make_response, current_app, Response
from flask.cli import AppGroup
from flask_babel import Babel,lazy_gettext as _, ngettext
from flask.json import jsonify
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from werkzeug.wsgi import wrap_file

from oajf import MESSAGE_TYPE_ERROR,MESSAGE_TYPE_SUCCESS,MESSAGE_TYPE_WARNING,MESSAGE_TYPE_INFO
from oajf.session import MariaDBSession
from oajf.db import get_db,DB,init as db_init
from oajf.db import (
    saveJournal as db_saveJournal, 
    deleteJournal as db_deleteJournal,
    readJournals as db_readJournals,
    readPublishers as db_readPublishers,
    savePublisher as db_savePublisher,
    deletePublisher as db_deletePublisher,
    readExcelFiles as db_readExcelFiles,
    saveExcelFile as db_saveExcelFile,
    deleteExcelFile as db_deleteExcelFile,
    readSettings as db_readSettings,
    saveSetting as db_saveSetting,
    deleteSetting as db_deleteSetting,
)
from oajf.models import OASTATUS, APPLICATION_REQUIREMENT, Journal, Publisher, Link, Excel, Setting, LINKTYPE, APPREQ_REQUIRED, APPREQ_NOT_REQUIRED
from oajf.util import logfunc,get_publishers,get_settings,getDOAJChangesFileAsExcelWorkbook,getDOAJDump,getSettingValueLang
from oajf.cli import register_cli

from oajf.config import LOGCONFIG
logging.config.dictConfig(LOGCONFIG)

app = Flask(__name__)
app.config.from_pyfile("oajf/config.py")
mailconfig = app.config.get('MAIL',None)
if mailconfig:
    to = mailconfig['to']
    if len(to) > 0:
        mailhandler = logging.handlers.SMTPHandler(
            mailhost=(mailconfig['host'], mailconfig['port']),
            fromaddr=mailconfig['from'],
            toaddrs=to,
            subject=mailconfig['subject'],
#                credentials=credentials,
#                secure=secure
        )
        mailhandler.setLevel(mailconfig['loglevel'])
        app.logger.addHandler(mailhandler)

#loggers = [logging.getLogger(name) for name in logging.root.manager.loggerDict]
#for logger in loggers:
#    print(logger)
#    print(logger.handlers)


# locale handling
def get_locale():
    lang = session.get("lang")
    if lang is not None:
        return lang
    session["lang"] = request.accept_languages.best_match(['de', 'en'])
    return session["lang"]

@app.route("/locale/<language>/<path>")
def set_locale(language=None,path=None):
    lang = session.get("lang")
    session["lang"] = language
    if path is None:
        return redirect(url_for('/'))
    else:
        return redirect(url_for(path))

app.config['TEMPLATES_AUTO_RELOAD'] = True
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

babel = Babel()
babel.init_app(app, locale_selector=get_locale)


app.logger.info(f"WERKZEUG_RUN_MAIN {os.environ.get("WERKZEUG_RUN_MAIN",'')}")
app.logger.info(f"SERVER_SOFTWARE {os.environ.get("SERVER_SOFTWARE",'')}")

# only setup after reload in debug mode (WERKZEUG), start immediately for gunicorn
if os.environ.get("WERKZEUG_RUN_MAIN",'') == "true":
    db = db_init(app)
    atexit.register(db.disconnect)
    db.signalhandler = signal.signal(signal.SIGINT,db.disconnect)

    MariaDBSession(app)

elif os.environ.get("SERVER_SOFTWARE",'').startswith("gunicorn"):
    db = db_init(app)
    atexit.register(db.disconnect)
    db.signalhandler = signal.signal(signal.SIGTERM,db.disconnect)

    MariaDBSession(app)

# TODO - check: 20250818
# make sure db gets inited even if none of the above conditions applies
# Werkzeug refused to start debug mode all of a sudden, started working again after
# recreating the virtual environment. huh?
else:
    db = db_init(app)

register_cli(app)

PAGE_LENGTH = 100

@app.context_processor
def utility_processor():
    return dict(
        CURRENT_LOCALE = get_locale(),
        OA_STATUS = OASTATUS,
        APPLICATION_REQUIREMENT = APPLICATION_REQUIREMENT,
        APPREQ_REQUIRED = APPREQ_REQUIRED,
        LINKTYPE = LINKTYPE,
        SUPERADMIN_UIDS =  current_app.config.get('SUPERADMIN_UIDS'),
        MESSAGE_TYPE_ERROR = MESSAGE_TYPE_ERROR,
        MESSAGE_TYPE_SUCCESS = MESSAGE_TYPE_SUCCESS,
        MESSAGE_TYPE_WARNING = MESSAGE_TYPE_WARNING,
        MESSAGE_TYPE_INFO = MESSAGE_TYPE_INFO,
        UI = current_app.config.get('UI'),
        getSettingValueLang = getSettingValueLang
        )

def login_required(view):
    @functools.wraps(view)
    def wrapped_view(**kwargs):
        if "uid" not in session:
            return redirect(url_for("admin_login"))
        return view(**kwargs)
    return wrapped_view

def superadmin_required(view):
    @functools.wraps(view)
    def wrapped_view(**kwargs):
        if "uid" not in session:
            return redirect(url_for("admin_login"))

        uid = session['uid']
        if 'SUPERADMIN_UIDS' not in app.config or uid not in app.config['SUPERADMIN_UIDS']:
            return redirect(url_for("admin_login"))

        return view(**kwargs)
    return wrapped_view

@app.route('/legal')
@logfunc
def legal():
    return render_template('legal.html')

@app.route('/data_protection')
@logfunc
def data_protection():
    return render_template('data_protection.html')

@app.route('/support')
@logfunc
def support():
    return render_template('support.html')

@app.route('/accessibility')
@logfunc
def accessibility():
    return render_template('accessibility.html')

@app.route("/", methods=('GET','POST'))
@logfunc
def index():
    get_publishers()
    if request.method == "GET":
        keyword = request.args.get("keyword",'')
        keyword = urllib.parse.unquote(keyword)
        keyword = urllib.parse.unquote_plus(keyword)
        page = request.args.get("page",None)
        order = request.args.get("order","")
        toggle = request.args.get("sort_toggle","")
    elif request.method == "POST":
        keyword = request.form['keyword']
        page = request.form.get("page",None)
        order = request.form.get("order","")
        toggle = request.form.get("sort_toggle","")
    else:
        keyword = ''
        page = None
        toggle = ''
        order = ''        

    if toggle:
        asc = toggle + '.ASC'
        desc = toggle + '.DESC'
        if toggle not in order:
            if order:
                order += ','
            order += asc
        else:
            if asc in order: order = order.replace(asc,desc)
            elif desc in order: order = order.replace(desc,'')
    order = order.strip(",")
    order = order.replace(",,",",")
    journals = db_readJournals(keyword=keyword,order=order)

    try:
        page = int(page)
    except:
        page = 0

    length = len(journals)
    number_of_pages = length // PAGE_LENGTH + 1
    page = min(max(page, 0), number_of_pages - 1)
    journals = journals[page*PAGE_LENGTH:((page+1)*PAGE_LENGTH)]
    start = (page * PAGE_LENGTH) + 1
    end = min((page + 1) * PAGE_LENGTH, length)
    return render_template("index.html", entries=journals, keyword=keyword, length=length, page=page,
                            pages=number_of_pages, start=start, end=end, order=order)

@app.post("/item_clicked")
@logfunc
def item_clicked():
    return ('', 204)

@app.post("/fetch")
@logfunc
def fetch_some_journals():
    get_publishers()
    keyword = request.form['keyword']
    journals = db_readJournals(keyword=keyword,
            order='title.ASC',
            limit=10,
            publisher_shallow=True,
            as_json=True)
    return journals

@app.route("/admin_login", methods=('GET', 'POST'))
def admin_login():
    if "uid" in session:
        return redirect(url_for("index"))
    if request.method == 'POST':
        uid = request.form['uid']
        password = request.form['password']
        if uid == "" or password == "":
            flash(_("Kein Passwort oder Benutzername angegeben."),MESSAGE_TYPE_ERROR)
            return render_template("admin_login.html")
        if authLDAP(uid, password):
            session["uid"] = uid
            return redirect(url_for("index"))
        else:
            flash(_('Falsches Passwort oder Benutzername.'),MESSAGE_TYPE_ERROR)

    return render_template("admin_login.html")


@app.route("/admin_logout")
@logfunc
@login_required
def admin_logout():
    session.clear()
    return redirect(url_for("index"))


@app.route("/admin_edit_journals", methods=('GET', 'POST'))
@logfunc
@login_required
def admin_edit_journals():
    publisher = None
    only_active = False
    page = 0
    toggle = ""
    keyword = ""
    order = ""
    action = "search"

    get_publishers()

    if request.method == 'GET': source = request.args
    elif request.method == 'POST': source = request.form
    else: source = None

    if source:
        keyword = source.get("keyword",None)
        only_active = source.get("only_active",None)
        publisher = source.get("publisher",None)
        toggle = source.get("sort_toggle","")
        order = source.get("order", "")
        page = source.get("page", 0, type=int)
        action = source.get("submit_action","search")
    
    only_active = True if only_active and only_active.lower() == 'on' else False
    if publisher: publisher = g.m_publishers.get(int(publisher),None)
    if toggle:
        asc = toggle + '.ASC'
        desc = toggle + '.DESC'
        if toggle not in order:
            if order:
                order += ','
            order += asc
        else:
            if asc in order: order = order.replace(asc,desc)
            elif desc in order: order = order.replace(desc,'')
    order = order.strip(",")
    order = order.replace(",,",",")

    journals = db_readJournals(keyword=keyword, only_active=only_active,publisher=publisher,order=order)

    if request.method == 'GET' or action == "search" or (request.method == 'POST' and 'btn-search' in request.form):
        length = len(journals)
        number_of_pages = length // PAGE_LENGTH + 1
        page = min(max(page, 0), number_of_pages - 1)
        journals = journals[page*PAGE_LENGTH:((page+1)*PAGE_LENGTH)]
        start = (page * PAGE_LENGTH) + 1
        end = min((page + 1) * PAGE_LENGTH, length)

        if keyword is None: keyword = ''

        return render_template("admin_edit_journals.html", entries=journals,
                            keyword=keyword, only_active=only_active, publisher=publisher,
                            page=page,
                            pages=number_of_pages, length=length, start=start, end=end, order=order)
    
    
    else:
        out = io.BytesIO()        
        wb = xlsxwriter.Workbook(out,{'in_memory': True})
        sheet = wb.add_worksheet('journals')
        row = 0
        col = iter(range(0,20))
        sheet.write(row,next(col),'id')
        sheet.write(row,next(col),'title')
        sheet.write(row,next(col),'url')
        sheet.write(row,next(col),'e-issn')
        sheet.write(row,next(col),'print-issn')
        sheet.write(row,next(col),'valid')
        sheet.write(row,next(col),'publisher')

        for j in journals:
            row += 1
            col = iter(range(0,20))
            sheet.write(row,next(col),j.id)
            sheet.write(row,next(col),j.title)
            sheet.write(row,next(col),j.url)
            sheet.write(row,next(col),j.e_issn)
            sheet.write(row,next(col),j.print_issn)
            sheet.write(row,next(col),str(j.valid_till))
            sheet.write(row,next(col),str(j.publisher))
        wb.close()

        # need to set to position 0
        out.seek(0)
        return send_file(out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Journals.xlsx')


@app.post("/admin_save_journal")
@logfunc
@login_required
def admin_save_journal():
    try:
        j = Journal()
        j.id = request.form["id"]
        j.title = request.form.get("title",None)
        j.url = request.form.get("url",None)
        j.print_issn = request.form.get("print_issn",None)
        j.e_issn = request.form.get("e_issn",None)
        j.valid_till = request.form.get("valid_till",None)
        db_saveJournal(j)
        flash(_('Zeitschrift gespeichert.'),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        flash("Speichern der Zeitschrift fehlgeschlagen.",MESSAGE_TYPE_ERROR)
        flash(str(e),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    # 307 redirect as post and keep form data
    return redirect(url_for('admin_edit_journals'),307)

@app.post("/admin_delete_journal")
@logfunc
@login_required
def admin_delete_journal():
    try:
        db_deleteJournal(None,id=request.form["journal_id_to_delete"])
        flash(_('Zeitschrift gelöscht.'),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        flash("Felher beim Löschen der Zeitschrift.",MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    # 307 redirect as post and keep form data
    return redirect(url_for('admin_edit_journals'),307)


@app.get("/admin_upload")
@logfunc
@login_required
def admin_upload_get():
    get_publishers()
    return render_template("admin_upload.html")


@app.post("/admin_upload")
@logfunc
@login_required
def admin_upload_post():
    get_publishers()
    conn: mariadb.Connection
    l_new: List[Journal] = []
    publisher: Publisher = None
    params = {}

    delete_journals = request.form.get('delete_journals',None)
    if "excel" not in request.files:
        flash(_("Kein File im request."),MESSAGE_TYPE_ERROR)
        return render_template("admin_upload.html",publisher=publisher)
    file = request.files["excel"]
    if file.filename == "":
        flash(_("Kein Excel hochgeladen."),MESSAGE_TYPE_ERROR)
        return render_template("admin_upload.html",publisher=publisher)
    if request.form["valid"] == "":
        flash(_("Kein Gültigkeitsende angegeben."),MESSAGE_TYPE_ERROR)
        return render_template("admin_upload.html",publisher=publisher)
    if request.form["publisher"] == "":
        flash(_("Kein Verlag angegeben."),MESSAGE_TYPE_ERROR)
        return render_template("admin_upload.html",publisher=publisher)
    
    if match := re.search(r'[\w\d_\-]+?\.xlsx',file.filename):
        pass
    else:
        flash(_("Falscher Dateityp, es werden nur .xlsx-Dateien unterstützt."),MESSAGE_TYPE_ERROR)
        return render_template("admin_upload.html",publisher=publisher)

    publisher_id = int(request.form["publisher"])
    publisher = g.m_publishers[publisher_id]
    valid = request.form['valid']
    valid_till =  datetime.datetime.strptime(request.form['valid'],'%Y-%m-%d').date()

    params['publisher'] = publisher
    params['valid'] = valid
    params['delete_journals'] = delete_journals

    filename = secure_filename(file.filename)
    file.seek(0)
    contents = file.read()
    file.seek(0)

    try:
        wb = openpyxl.load_workbook(file, read_only=True,data_only=True)
        ws = wb.active

        if (ws.cell(1,1).value.lower() != 'titel' and ws.cell(1,1).value.lower() != 'title'):
            flash(_("'Titel' oder 'Title' in Zelle A1 erwartet."),MESSAGE_TYPE_ERROR)
            return render_template("admin_upload.html",**params)
        if (ws.cell(1,2).value.lower() != 'link'):
            flash(_("'Link' in Zelle B1 erwartet."),MESSAGE_TYPE_ERROR)
            return render_template("admin_upload.html",**params)
        if (ws.cell(1,3).value.lower() != 'e-issn'):
            flash(_("'E-ISSN' in Zelle C1 erwartet."),MESSAGE_TYPE_ERROR)
            return render_template("admin_upload.html",**params)
        if (ws.cell(1,4).value.lower() != 'print-issn'):
            flash(_("'Print-ISSN' in Zelle D1 erwartet."),MESSAGE_TYPE_ERROR)
            return render_template("admin_upload.html",**params)

        row_idx = 1
        content_errors = 0
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            if content_errors > 4:
                break

            if row[0] is None:
                break

            row_idx += 1
            
            title = str(row[0])
            title = title.strip()
            link = str(row[1])
            link = link.strip()

            e_issn = str(row[2])
            e_issn = e_issn.strip()
            if e_issn and e_issn != 'None':
                if m:=re.search('([0-9X]{4}-[0-9X]{4})',e_issn): 
                    e_issn = m[1]
                else:
                    content_errors += 1
                    msg = _(f"Ungültige E-ISSN {0} (Zeile {1})") 
                    flash(msg.format(e_issn,row_idx),MESSAGE_TYPE_ERROR)

            print_issn = str(row[3])
            print_issn = print_issn.strip()
            if print_issn and print_issn != 'None':
                if m:=re.search('([0-9X]{4}-[0-9X]{4})',print_issn): 
                    print_issn = m[1]
                else:
                    content_errors += 1
                    msg = _(f"Ungültige Print-ISSN {0} (Zeile {1})") 
                    flash(msg.format(print_issn,row_idx),MESSAGE_TYPE_ERROR)

            if e_issn == 'None': e_issn = None
            if print_issn == 'None': print_issn = None
            if link == 'None': link = None

            j = Journal()
            j.title = title
            j.publisher = publisher
            j.valid_till = valid_till
            j.title = title
            j.url = link
            j.e_issn = e_issn
            j.print_issn = print_issn
            l_new.append(j)

    except Exception as e:
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
        flash(_("Import fehlgeschlagen. Fehler beim Parsen der Input-Datei."),MESSAGE_TYPE_ERROR)
        return render_template("admin_upload.html",**params)

    if content_errors > 0:
        return render_template("admin_upload.html",**params)

    try:
        conn = get_db()
        if delete_journals:
            cnt_deleted_journals = db_deleteJournal(None,transaction_conn=conn,publisher_id=publisher_id)

        e = Excel()
        e.name = filename
        e.file = contents
        e.valid = valid
        e.publisher = g.m_publishers[publisher_id]

        db_saveExcelFile(e,transaction_conn=conn)
    
        for j in l_new:
            db_saveJournal(j,conn)

        conn.commit()
        flash(_("Excel-Datei erfolgreich importiert."),MESSAGE_TYPE_SUCCESS)
        msg = ngettext("{0} Zeitschrift gelöscht.","{0} Zeitschriften gelöscht.",cnt_deleted_journals)
        flash(msg.format(cnt_deleted_journals),MESSAGE_TYPE_SUCCESS)
        
        flash(ngettext(f"{len(l_new)} Zeitschriften importiert",f"{len(l_new)} Zeitschriften importiert",len(l_new)),MESSAGE_TYPE_SUCCESS)

    except Exception as e:
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
        flash("Import aufgrund eines Datenbankfehlers fehlgeschlagen",MESSAGE_TYPE_ERROR)
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()
    
    return render_template("admin_upload.html",**params)


@app.route("/admin_delete", methods = ['GET','POST'])
@logfunc
@login_required
def admin_delete():
    get_publishers()
    if request.method == 'GET':
        return render_template("admin_delete.html")

    conn: mariadb.Connection
    cnt_deleted = 0

    if "excel" not in request.files:
        flash("Keine Datei hochgeladen.",MESSAGE_TYPE_ERROR)
        return render_template("admin_delete.html")
    file = request.files["excel"]
    print (file.filename)
    if match := re.search(r'[\w\d_\-]+?\.xlsx',file.filename):
        pass
    else:
        flash("Falscher Dateityp, es werden nur .xlsx-Dateien unterstützt",MESSAGE_TYPE_ERROR)
        return render_template("admin_delete.html")

    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active
    mode = ws.cell(1,1).value
    if mode is None:
        flash("Spaltenkopf darf nicht leer sein.",MESSAGE_TYPE_ERROR)
        return render_template("admin_delete.html")
    
    mode = mode.lower()
    if mode != 'e-issn' and mode != 'id':
        flash("Spaltenkopf muss 'E-ISSN' oder 'Id' sein.",MESSAGE_TYPE_ERROR)
        return render_template("admin_delete.html")
        

    try:
        conn = get_db()
        cur = conn.cursor()

        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is None:
                break
            value = str(row[0])
            if value:
                value = value.strip()
                if mode == 'e-issn':
                    if value and (m:=re.search('([0-9X]{4}-[0-9X]{4})',value)): 
                        value = m[1]
                    else:
                        value = None
                else:
                    try:
                        value = int(value)
                    except:
                        value = None

            if value:
                if mode == 'e-issn':
                    cnt_deleted += db_deleteJournal(None,transaction_conn=conn,e_issn=value)
                else:
                    cnt_deleted += db_deleteJournal(None,transaction_conn=conn,id=value)
        conn.commit()
        flash(f"{cnt_deleted} Zeitschriften gelöscht.",MESSAGE_TYPE_SUCCESS)            
    except Exception as e:
        flash( "Löschen der Zeitschriften fehlgeschlagen.",MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()

    return render_template("admin_delete.html")


#
# settings
#
@app.get("/admin_settings")
@logfunc
@login_required
def admin_settings():
    get_settings()
    return render_template("admin_settings.html")

@app.post("/admin_delete_setting")
@logfunc
@login_required
def admin_delete_setting():
    id = int(request.form["id"])
    try:
        db_deleteSetting(None,id=id)
        flash(_("Einstellung gelöscht."),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        flash(_('Fehler beim Löschen der Einstellung.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    get_settings(force_reload=True)
    return redirect(url_for("admin_settings"))

@app.post("/admin_copy_setting")
@logfunc
@login_required
def admin_copy_setting():
    try:
        get_settings()
        id = int(request.form["id"])
        o = g.m_id_setting[id]
        o.id = -1
        o.name = o.name + ' (Kopie)'
        o = db_saveSetting(o)
        flash(_('Einstellung kopiert.'),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        flash(_('Fehler beim Kopieren der Einstellung.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    get_settings(force_reload=True)
    return redirect(url_for('admin_settings'))

@app.post("/admin_save_setting")
@logfunc
@login_required
def admin_save_setting():
    try:
        o = Setting()
        o.id = int(request.form["id"])
        o.name = request.form.get("name",None)
        o.value = request.form.get("value",None)
        o.value_en = request.form.get("value_en",None)
        o.value_de = request.form.get("value_de",None)
        o = db_saveSetting(o)

        flash(_('Einstellung gespeichert.'),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        flash(_('Fehler beim Speichern der Einstellung.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    get_publishers(force_reload=True)
    return redirect(url_for('admin_settings'))

#
# publisher
#
@app.get("/admin_publishers")
@logfunc
@login_required
def admin_publishers_get():
    get_publishers()
    return render_template("admin_publishers.html")

@app.get("/admin_export_publishers_as_json")
@logfunc
@login_required
def admin_export_publishers_as_json():
    p: Publisher
    l: List[Dict] = []
    get_publishers()
    for p in g.publishers:
        d = p.toDict(includeid=False)
        l.append(d)

    content = json.dumps(l,indent=4)
    return Response(content, 
            mimetype='application/json',
            headers={'Content-Disposition':'attachment;filename=publishers.json'})


@app.post("/admin_delete_publisher")
@logfunc
@login_required
def admin_delete_publisher():
    id = int(request.form["id"])
    try:
        conn = get_db()
        id = int(request.form["id"])
        cnt_deleted_journal = db_deleteJournal(None,transaction_conn=conn,publisher_id=id)
        cnt_deleted_excel = db_deleteExcelFile(None,transaction_conn=conn,publisher_id=id)
        db_deletePublisher(None,transaction_conn=conn,id=id)
        conn.commit()
        msg = ngettext('{0} Zeitschrift gelöscht', '{0} Zeitschriften gelöscht.', cnt_deleted_journal)
        flash(msg.format(cnt_deleted_journal),MESSAGE_TYPE_SUCCESS)
        msg = ngettext('{0} Excel-File gelöscht', '{0} Excel-Files gelöscht.', cnt_deleted_excel)
        flash(msg.format(cnt_deleted_excel),MESSAGE_TYPE_SUCCESS)
        flash(_("Verlag gelöscht."),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        if conn:
            conn.rollback()
        flash(_('Fehler beim Löschen des Verlags.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
    finally:
        if conn:
            conn.close()

    get_publishers(force_reload=True)
    return redirect(url_for("admin_publishers_get"))

@app.post("/admin_copy_publisher")
@logfunc
@login_required
def admin_copy_publisher():
    try:
        get_publishers()
        id = int(request.form["id"])
        p = g.m_publishers[id]
        # potentially unsafe
        p.id = -1
        p.name = p.name + ' (Kopie)'
        p.is_doaj = 0
        p = db_savePublisher(p)
        flash(_('Verlag kopiert.'),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        flash(_('Fehler beim Kopieren des Verlags.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    get_publishers(force_reload=True)
    return redirect(url_for('admin_publishers_get'))

@app.post("/admin_save_publisher")
@logfunc
@login_required
def admin_save_publisher():
    try:
        p = Publisher()
        p.id = int(request.form["id"])
        p.name = request.form.get("name",None)
        p.validity = request.form.get("validity",None)
        p.oa_status = OASTATUS.get(request.form.get("oa_status",None),None)
        p.application_requirement = APPLICATION_REQUIREMENT.get(request.form.get("application_requirement",None),None)
        p.funder_info = request.form.get("funder_info",None)
        p.cost_coverage = request.form.get("cost_coverage",None)
        p.valid_tu = request.form.get("valid_tu",None)
        p.article_type = request.form.get("article_type",None)
        p.further_info = request.form.get("further_info",None)
        p.funder_info_en = request.form.get("funder_info_en",None)
        p.cost_coverage_en = request.form.get("cost_coverage_en",None)
        p.valid_tu_en = request.form.get("valid_tu_en",None)
        p.article_type_en = request.form.get("article_type_en",None)
        p.further_info_en = request.form.get("further_info_en",None)
        p.is_doaj = request.form.get("is_doaj",None)
        p.is_doaj = 1 if p.is_doaj and p.is_doaj == 'on' else 0
        p.doaj_linked = request.form.get("doaj_linked",None)
        p.doaj_linked = 1 if p.doaj_linked and p.doaj_linked == 'on' else 0
        
        # doesn't work like this for checkboxes !! since checkboxes not included if unchecked
        links = []
        links_id = request.form.getlist('links_id[]')
        links_linktype = request.form.getlist('links_linktype[]')
        links_link = request.form.getlist('links_link[]')
        links_linktext_de = request.form.getlist('links_linktext_de[]')
        links_linktext_en = request.form.getlist('links_linktext_en[]')

        for i,x in enumerate(links_id):
            link = Link()
            link.link = links_link[i]
            link.linktype = LINKTYPE.get(links_linktype[i],None)
            link.linktext_de = links_linktext_de[i]
            link.linktext_en = links_linktext_en[i]
            links.append(link)
        
        p.links = links
        p = db_savePublisher(p)

        flash(_('Verlag gespeichert.'),MESSAGE_TYPE_SUCCESS)
    except Exception as e:
        flash(_('Fehler beim Speichern des Verlags.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    get_publishers(force_reload=True)
    return redirect(url_for('admin_publishers_get'))

@app.get("/admin_excel_list")
@logfunc
@login_required
def admin_excel_list():
    l_excel: List[Excel] = []

    try:
        l_excel = db_readExcelFiles()
    except Exception as e:
        flash(_('Fehler beim Lesen der Excel-Files.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    return render_template("admin_excel_list.html",l_excel=l_excel)

@app.get('/admin_excel_download/<path:id>')
@logfunc
@login_required
def admin_excel_download(id):
    l_excel: List[Excel] = [] 
    e: Excel = None

    try:
        l_excel: List[Excel] = db_readExcelFiles(id=id,include_data=True)
    except Exception as e:
        flash(_('Fehler beim Lesen der Excel-Files.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
        return render_template("admin_excel_list.html",l_excel=l_excel)

    # invalid id
    if not l_excel:
        flash(_('Fehler beim Lesen des Excel-Files.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")
        return render_template("admin_excel_list.html",l_excel=l_excel)

    e = l_excel[0]
    return send_file(
        io.BytesIO(e.file),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=e.name)

@app.post('/admin_excel_delete')
@logfunc
@login_required
def admin_excel_delete():
    try:
        id = request.form.get("id",None)
        if id:
            db_deleteExcelFile(None,id=id)
            flash(_("Excel-File gelöscht."),MESSAGE_TYPE_SUCCESS)
        else:
            flash(_('File-Id fehlt im Aufruf.'),MESSAGE_TYPE_WARNING)
    except Exception as e:
        flash(_('Fehler beim Löschen des Excel-Files.'),MESSAGE_TYPE_ERROR)
        app.logger.error(f"exception={type(e).__name__}")
        app.logger.error(f"stacktrace={traceback.format_exc()}")

    return redirect(url_for('admin_excel_list'))


@app.route("/doaj_import_update",methods=['GET','POST'])
@logfunc
@login_required
def doaj_import_update():
    l_new: List[Journal] = []
    l_updated: List[Journal] = []
    j: Journal = None
    p: Publisher = None

    publishers = get_publishers()
    publishers[:] = [p for p in publishers if p.is_doaj == 1]
    p = publishers[0]

    if request.method == 'GET':
        try:
            l_journal_csv: List[Journal] = []
            l_journal_db: List[Journal] 
            m_eissn: Dict[str,Journal] = {}
            m_pissn: Dict[str,Journal] = {}

            l_journal_db = db_readJournals(publisher=p,publisher_shallow=True)

            for j in l_journal_db:
                if j.e_issn:
                    if j.e_issn in m_eissn:
                        flash(_(f"E-ISSN mehrfach gefunden für Zeitschriften in der Datenbank: {j.e_issn}"))
                    else:
                        m_eissn[j.e_issn] = j

                if j.print_issn:
                    if j.print_issn in m_pissn:
                        flash(_(f"Print-ISSN mehrfach gefunden für Zeitschriften in der Datenbank: {j.print_issn}"))
                    else:
                        m_pissn[j.print_issn] = j

            l_journal_csv,errs = getDOAJDump()
            if errs:
                for e in errs:
                    flash(e,MESSAGE_TYPE_ERROR)
                return render_template("admin_doaj_dump.html",l_new=[],l_updated=[])

            for j in l_journal_csv:
                e_issn_found = False
                print_issn_found = False
                if j.e_issn:
                    if j.e_issn in m_eissn:
                        e_issn_found = True
            
                if j.print_issn:
                    if j.print_issn in m_pissn:
                        print_issn_found = True

                if e_issn_found:
                    j_db = m_eissn[j.e_issn]
                    diffs = j.getDifferences(j_db)
                    if diffs:
                        j.id = j_db.id
                        j.valid_till = j_db.valid_till
                        j.diffs = diffs
                        l_updated.append(j)
                elif print_issn_found:
                    j_db = m_pissn[j.print_issn]
                    diffs = j.getDifferences(j_db)
                    if diffs:
                        j.id = j_db.id
                        j.valid_till = j_db.valid_till
                        j.diffs = diffs
                        l_updated.append(j)
                else:
                    j.id = -1
                    j.publisher = p
                    j.valid_till = datetime.datetime.today()
                    j.valid_till = j.valid_till.replace(month=12,day=31,hour=23,minute=59,second=59)
                    l_new.append(j)
            
            if len(l_new) == 0 and len(l_updated) == 0:
                flash("Weder neue noch zu aktualisierende Zeitschriften gefunden.",MESSAGE_TYPE_WARNING)

        except Exception as e:
            l_new = []
            l_updated = []
            app.logger.error(f"exception={type(e).__name__}")
            app.logger.error(f"stacktrace={traceback.format_exc()}")
            flash('Fehler beim Holen der DOAJ-Änderungen.',MESSAGE_TYPE_ERROR)
            flash(e,MESSAGE_TYPE_ERROR)

        return render_template("admin_doaj_dump.html",l_new=l_new,l_updated=l_updated)
    elif request.method == 'POST':
        data = request.form.get("data_as_json")
        data = json.loads(data)
        action = data.get('action')
        if action == 'import':
            try:
                conn = get_db()
                l = data.get('journals',[])
                for a in l:
                    j = Journal()
                    j.id = -1
                    j.publisher = p
                    j.valid_till = datetime.datetime.today()
                    j.valid_till = j.valid_till.replace(month=12,day=31,hour=23,minute=59,second=59)
                    j.title = a['title']
                    j.e_issn = a['e_issn']
                    j.print_issn = a['print_issn']
                    j.url = a['url']
                    l_new.append(j)

                for j in l_new:
                    j = db_saveJournal(j,transaction_conn=conn)

                conn.commit()

                msg = ngettext("{0} Zeitschrift importiert.","{0} Zeitschriften importiert.",len(l_new) )
                flash(msg.format(len(l_new)),MESSAGE_TYPE_SUCCESS)
            except Exception as e:
                if conn is not None:
                    conn.rollback()
                app.logger.error(f"exception={type(e).__name__}")
                app.logger.error(f"stacktrace={traceback.format_exc()}")
                flash(_('Fehler beim Import der Zeitschriften.'),MESSAGE_TYPE_ERROR)
                flash(e,MESSAGE_TYPE_ERROR)
            finally:
                if conn is not None:
                    conn.close()
        elif action == 'update':
            try:
                conn = get_db()
                l = data.get('journals',[])
                for a in l:
                    j = db_readJournals(transaction_conn=conn,id = int(a['id']))[0]
                    j.title = a['title']
                    j.e_issn = a['e_issn']
                    j.print_issn = a['print_issn']
                    j.url = a['url']
                    l_updated.append(j)

                for j in l_updated:
                    j = db_saveJournal(j,transaction_conn=conn)

                conn.commit()

                msg = ngettext("{0} Zeitschrift aktualisiert.","{0} Zeitschriften aktualisiert.",len(l_updated) )
                flash(msg.format(len(l_updated)),MESSAGE_TYPE_SUCCESS)
            except Exception as e:
                if conn is not None:
                    conn.rollback()
                app.logger.error(f"exception={type(e).__name__}")
                app.logger.error(f"stacktrace={traceback.format_exc()}")
                flash('Fehler beim Aktualisieren der Zeitschriften.',MESSAGE_TYPE_ERROR)
                flash(e,MESSAGE_TYPE_ERROR)
            finally:
                if conn is not None:
                    conn.close()

        return render_template("admin_doaj_dump.html",l_new=[],l_updated=[])

@app.route("/doaj_withdrawn",methods=['GET','POST'])
@logfunc
@login_required
def doaj_withdrawn():
    if request.method == 'GET':
        map_issn: Dict[str,List[str]] = {}
        l_journal: List[Journal] = []

        wb,data,errs = getDOAJChangesFileAsExcelWorkbook()
        if errs:
            for e in errs:
                flash(e,MESSAGE_TYPE_ERROR)
            return render_template("admin_doaj_withdrawn.html",l_journal=[])

        try:
            sheet = wb['Withdrawn']
            for i,row in enumerate(sheet.rows):
                if i < 6:
                    continue

                title = row[0].value 
                if title: title = title.strip()
                issn = row[1].value 
                if issn: issn = issn.strip()
                date = row[2].value
                reason = row[3].value 
                if reason: reason = reason.strip()

                if issn is not None and issn != 'None':
                    map_issn[issn] = [title,date,reason]

            get_publishers()

            for k,v in map_issn.items():
                for j in db_readJournals(e_issn=k,only_active=True,publisher_shallow=True):
                    j.publisher = g.m_publishers[j.publisher.id]
                    j.withdraw_reason = map_issn[k][2]
                    j.withdraw_date = map_issn[k][1]
                    if j.publisher.is_doaj == 1 or j.publisher.doaj_linked == 1:
                        j.to_be_deleted = 1 
                    else:
                        j.to_be_deleted = 0

                    l_journal.append(j)
            
            if len(l_journal) == 0:
                flash("Keine zu löschenden Zeitschriften gefunden.",MESSAGE_TYPE_WARNING)

        except Exception as e:
            l_journal = []
            app.logger.error(f"exception={type(e).__name__}")
            app.logger.error(f"stacktrace={traceback.format_exc()}")
            flash('Fehler beim Holen der DOAJ-Änderungen.',MESSAGE_TYPE_ERROR)
            flash(e,MESSAGE_TYPE_ERROR)

        return render_template("admin_doaj_withdrawn.html",l_journal=l_journal)
    elif request.method == 'POST':
        ids = []

        s = request.form.get('ids','')
        s = s.strip()
        if s:
            ids = s.split(',')
            ids[:] = [int(id.strip()) for id in ids]

        try:
            conn = get_db()
            cur = conn.cursor()
            for id in ids:
                db_deleteJournal(None,transaction_conn=conn,id=id)
            conn.commit()


            flash(f"{len(ids)} Zeitschriften gelöscht",MESSAGE_TYPE_SUCCESS)
        except Exception as e:
            if conn is not None:
                conn.rollback()
            app.logger.error(f"exception={type(e).__name__}")
            app.logger.error(f"stacktrace={traceback.format_exc()}")
            flash('Fehler beim Löschen der Zeitschriften.',MESSAGE_TYPE_ERROR)
            flash(e,MESSAGE_TYPE_ERROR)
        finally:
            if conn is not None:
                conn.close()

        return render_template("admin_doaj_withdrawn.html",l_journal=[])






# --------------------------------------------------------------------------------------
def authLDAP(uid: str, password: str):
    """
    authenticates user against LDAP server
    optionally only gets user data for work as mode
    """

    def checkPassword(challenge_password, given_password):
        challenge_bytes = decode(challenge_password[6:])
        digest = challenge_bytes[:20]
        salt = challenge_bytes[20:]
        hr = hashlib.sha1(given_password)
        hr.update(salt)
        x = hr.digest()
        return digest == x

    try:
        server = Server(app.config["LDAP"]["server"], use_ssl=True, get_info=ALL)
        conn = Connection(server, 'cn=' + app.config["LDAP"]["bind_user"] + ',' + app.config["LDAP"]["search_base"],
                          password=app.config["LDAP"]["bind_user_password"], auto_bind=True, raise_exceptions=True)
        success = conn.search(
            search_base=app.config["LDAP"]["search_base"],
            search_filter='(uid=' + uid + ')',
            search_scope='SUBTREE',
            attributes=['cn', 'uid', 'userPassword', 'serviceName']
        )

        if success:
            app.logger.info(f"user {uid} found in LDAP")
        else:
            app.logger.info(f"user {uid} not found in LDAP")
            return False

    except Exception as e:
        app.logger.error(f"exception={type(e).__name__}")
        return False

    if conn.response:
        entry = conn.response[0]
        pw_hashed = entry['attributes']['userPassword'][0]
        uid = entry['attributes']['uid'][0] if len(entry['attributes']['uid']) > 0 else ''
        if app.config["LDAP"]["service_name"] in entry['attributes']['serviceName'] and checkPassword(pw_hashed, str.encode(password)):
            app.logger.info(f"correct password for user {uid} in LDAP")
            return True
        else:
            app.logger.info(f"incorrect password for user {uid} in LDAP")
            return False
    else:
        return False

with app.test_request_context():
    pass
